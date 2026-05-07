"""
Módulo: Relatório de Cadeiras (Explainable AI / Auditoria Visual)
Objetivo: Gerar arquivos de saída com expansão de candidatos por AMBIGUIDADE.
Novidade V7.7: INJEÇÃO DE AUDITORIA. Lê os "P" do arquivo humano e altera
               matematicamente o cálculo de Absenteísmo e a Matriz Final (Ouro).
"""

import os
import re
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from core import config_ambiente

def gerar_relatorio_cadeiras():
    print("==========================================================")
    print("📊 INICIANDO GERAÇÃO DOS ARQUIVOS DE AUDITORIA (XAI - V7.7)")
    print("==========================================================\n")

    if not os.path.exists(config_ambiente.CAMINHO_CSV_PRESENCA) or not os.path.exists(config_ambiente.CAMINHO_CSV_VISITANTES):
        print("❌ ERRO: Bases de dados não encontradas.")
        return

    print("⏳ Lendo bases (Camada Bronze) e processando metadados...")
    df_pres = pd.read_csv(config_ambiente.CAMINHO_CSV_PRESENCA, sep=';')
    df_vis = pd.read_csv(config_ambiente.CAMINHO_CSV_VISITANTES, sep=';')

    df_pres['Data'] = df_pres['Data'].astype(str).str.strip()
    df_vis['Data'] = df_vis['Data'].astype(str).str.strip()

    df_pres['Data_DT'] = pd.to_datetime(df_pres['Data'], format='%d/%m/%Y', errors='coerce')
    if df_pres['Data_DT'].isnull().all():
        df_pres['Data_DT'] = pd.to_datetime(df_pres['Data'], errors='coerce')

    regras = config_ambiente.REGRAS_CONSELHO
    termo_extra = regras.get("termo_reuniao_extra", "Extraordinária")
    regex_extra = regras.get("regex_arquivo_extra", r"extra_(\d+)")
    label_ia = regras.get("label_auditoria_ia", "Possível Conselheiro")

    if 'Reuniao' in df_pres.columns:
        def formatar_ata_v2(reuniao_str):
            reuniao_str = str(reuniao_str)
            match = re.search(r'(\d+)', reuniao_str)
            if match:
                return f"{match.group(1).zfill(2)}ª"
            return reuniao_str.split()[0].capitalize()

        df_pres['Numero_Ata'] = df_pres['Reuniao'].apply(formatar_ata_v2)

        if 'Arquivo' in df_pres.columns:
            masc_extra = df_pres['Reuniao'].astype(str).str.contains(termo_extra, case=False, na=False)
            num_extra = df_pres.loc[masc_extra, 'Arquivo'].astype(str).str.extract(regex_extra)[0]
            df_pres.loc[masc_extra, 'Numero_Ata'] = "Extra " + num_extra.str.zfill(2).fillna('?')

        df_pres['Numero_Ata'] = df_pres['Numero_Ata'].fillna('-')
        df_pres['Ref_Reuniao'] = df_pres['Numero_Ata'] + " - " + df_pres['Data']
    else:
        df_pres['Ref_Reuniao'] = df_pres['Data']

    # ========================================================
    # 🎯 LEITURA ANTECIPADA DA MEMÓRIA & EXTRAÇÃO DE PRESENÇAS
    # ========================================================
    path_auditoria = os.path.join(config_ambiente.CAMINHO_PROCESSADOS, "Auditoria_Humana_XAI.xlsx")
    col_conf = 'Conferencia_Humana (Digite: P=Presente, V=Visitante ou -=Descarte)'

    fragmentos_resolvidos = set()
    presencas_auditadas = set() # Guarda as cadeiras que ganharam "P"
    df_antigo = None

    if os.path.exists(path_auditoria):
        print("🔍 Analisando decisões humanas anteriores para injetar na Matriz...")
        df_antigo = pd.read_excel(path_auditoria, sheet_name='Tarefas_Auditoria')

        if 'Status_Sugerido' in df_antigo.columns:
            df_antigo = df_antigo.drop(columns=['Status_Sugerido'])

        if 'Data_DT' not in df_antigo.columns:
            df_antigo['Data_DT'] = pd.to_datetime(df_antigo['Reuniao_Referencia'].str.split(' - ').str[-1], errors='coerce')

        if 'Nome_na_Ata (O que a IA leu)' in df_antigo.columns:
            df_antigo['Nome_na_Ata (O que a IA leu)'] = df_antigo['Nome_na_Ata (O que a IA leu)'].astype(str).str.title()

        for _, r in df_antigo.iterrows():
            decisao = str(r.get(col_conf, '')).strip().upper()
            ref_reuniao = r['Reuniao_Referencia']
            frag = str(r.get('Nome_na_Ata (O que a IA leu)', ''))

            if decisao in ['P', 'V', '-']:
                fragmentos_resolvidos.add((ref_reuniao, frag))

            # Se você deu "P", guardamos a Cadeira para forçar a presença na matemática
            if decisao == 'P':
                chave_presenca = (r['Segmento'], r['Cadeira_Original'], ref_reuniao)
                presencas_auditadas.add(chave_presenca)

    # ========================================================
    # MAPEAMENTO DE FRAGMENTOS LIDOS (IA)
    # ========================================================
    mapa_fragmentos_ia = {}
    filtro_ia = df_vis['Tipo_Visitante'].str.contains(label_ia, case=False, na=False)

    for _, v_row in df_vis[filtro_ia].iterrows():
        data_ata = str(v_row['Data']).strip()
        texto_ocr = str(v_row.get('Nome_na_Ata', '')).strip()
        if data_ata not in mapa_fragmentos_ia:
            mapa_fragmentos_ia[data_ata] = {}
        if len(texto_ocr) > 2:
            mapa_fragmentos_ia[data_ata][texto_ocr.upper()] = texto_ocr

    lista_aba_auditoria = []
    set_cadeira_pendente = set()

    print("🧩 Expandindo candidatos...")
    for _, row in df_pres.iterrows():
        if row['Presente'] == 0:
            nome_oficial = str(row['Nome']).strip()
            nome_upper = nome_oficial.upper()
            data_bruta = str(row['Data']).strip()

            fragments_for_date = mapa_fragmentos_ia.get(data_bruta, {})

            for f_upper, f_orig in fragments_for_date.items():
                f_title = str(f_orig).title()
                reuniao_ref = row['Ref_Reuniao']

                if (reuniao_ref, f_title) in fragmentos_resolvidos:
                    continue

                words_f = set(f_upper.split())
                words_n = set(nome_upper.split())

                if words_f.intersection(words_n):
                    chave_info = (row['Segmento'], row['Cadeira'], reuniao_ref)
                    set_cadeira_pendente.add(chave_info)

                    lista_aba_auditoria.append({
                        'Reuniao_Referencia': reuniao_ref,
                        'Data_DT': row['Data_DT'],
                        'Segmento': row['Segmento'],
                        'Orgao_Representado': row.get('Orgao', 'N/A'),
                        'Cadeira_Original': row['Cadeira'],
                        'Nome_do_Conselheiro': nome_oficial,
                        'Nome_na_Ata (O que a IA leu)': f_title,
                        'Tipo_Vinculo': str(row['Tipo']).strip(),
                        'Conferencia_Humana (Digite: P=Presente, V=Visitante ou -=Descarte)': ''
                    })

    # ========================================================
    # 🎯 INJEÇÃO MATEMÁTICA & CONSOLIDAÇÃO DA MATRIZ (Snapshot)
    # ========================================================
    print("📈 Aplicando presenças humanas e consolidando a matriz...")
    df_cadeira_reuniao = df_pres.groupby(['Segmento', 'Cadeira', 'Ref_Reuniao', 'Data_DT'])['Presente'].max().reset_index()

    # MÁGICA DA V7.7: Transforma a falta em presença se você marcou 'P' na auditoria
    def aplicar_auditoria_humana(row):
        if row['Presente'] == 0:
            chave = (row['Segmento'], row['Cadeira'], row['Ref_Reuniao'])
            if chave in presencas_auditadas:
                return 1 # Força a presença!
        return row['Presente']

    df_cadeira_reuniao['Presente'] = df_cadeira_reuniao.apply(aplicar_auditoria_humana, axis=1)

    # Agora a estatística será 100% fiel ao seu trabalho humano
    estat = df_cadeira_reuniao.groupby(['Segmento', 'Cadeira']).agg(Total=('Ref_Reuniao','count'), Pres=('Presente','sum')).reset_index()
    estat['Faltas'] = estat['Total'] - estat['Pres']
    estat['% Abs'] = (estat['Faltas'] / estat['Total'] * 100).round(1)

    matriz = df_cadeira_reuniao.pivot_table(index=['Segmento', 'Cadeira'], columns='Ref_Reuniao', values='Presente', aggfunc='max').fillna('N/A')
    cols_ord = df_cadeira_reuniao.sort_values('Data_DT')['Ref_Reuniao'].unique()
    matriz = matriz.reindex(columns=cols_ord).replace({1: 'Presente', 0: 'Falta'})

    # Só marca como pendente se continuar a ser falta E estiver na lista de dúvidas
    for idx in matriz.index:
        for c in matriz.columns:
            if matriz.at[idx, c] == 'Falta' and (idx[0], idx[1], c) in set_cadeira_pendente:
                matriz.at[idx, c] = '⚠️ PENDENTE'

    matriz_final = pd.merge(estat, matriz.reset_index(), on=['Segmento', 'Cadeira'], how='left')

    # ========================================================
    # 🎯 CONSOLIDAÇÃO FINAL E LIMPEZA
    # ========================================================
    print("🧩 Consolidando tarefas finais...")
    df_novas_tarefas = pd.DataFrame(lista_aba_auditoria)

    if df_antigo is not None:
        df_consolidado = pd.concat([df_antigo, df_novas_tarefas], ignore_index=True)
        df_consolidado[col_conf] = df_consolidado[col_conf].replace('', pd.NA)
        df_consolidado = df_consolidado.sort_values(by=col_conf, na_position='last')

        df_lista_auditoria = df_consolidado.drop_duplicates(
            subset=['Reuniao_Referencia', 'Cadeira_Original', 'Nome_do_Conselheiro', 'Nome_na_Ata (O que a IA leu)'],
            keep='first'
        ).copy()

        df_lista_auditoria[col_conf] = df_lista_auditoria[col_conf].fillna('')

        def eh_obsoleto(r):
            esta_vazio = str(r[col_conf]).strip() == ''
            ja_resolvido = (r['Reuniao_Referencia'], str(r['Nome_na_Ata (O que a IA leu)'])) in fragmentos_resolvidos
            return esta_vazio and ja_resolvido

        mascara_obsoleto = df_lista_auditoria.apply(eh_obsoleto, axis=1)
        df_lista_auditoria = df_lista_auditoria[~mascara_obsoleto]

    else:
        df_lista_auditoria = df_novas_tarefas

    if not df_lista_auditoria.empty:
        df_lista_auditoria = df_lista_auditoria.sort_values(by=['Data_DT', 'Nome_na_Ata (O que a IA leu)', 'Nome_do_Conselheiro'])
        df_lista_auditoria = df_lista_auditoria.drop(columns=['Data_DT'], errors='ignore')

    # ========================================================
    # EXPORTAÇÃO COM FORMATAÇÃO (Openpyxl)
    # ========================================================
    path_matriz = config_ambiente.CAMINHO_EXCEL_CADEIRAS
    os.makedirs(os.path.dirname(path_matriz), exist_ok=True)
    os.makedirs(os.path.dirname(path_auditoria), exist_ok=True)

    f_pnd = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    f_pre = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    f_fal = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    f_inp = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    font_pnd = Font(color="9C6500", bold=True)

    with pd.ExcelWriter(path_matriz, engine='openpyxl') as writer:
        matriz_final.to_excel(writer, sheet_name='Matriz_Snapshot_IA', index=False)
        ws1 = writer.sheets['Matriz_Snapshot_IA']
        ws1.freeze_panes = "G2"
        for row in ws1.iter_rows(min_row=2, min_col=7):
            for cell in row:
                val = str(cell.value)
                if 'Presente' in val: cell.fill = f_pre
                elif 'Falta' in val: cell.fill = f_fal
                elif '⚠️' in val: cell.fill = f_pnd; cell.font = font_pnd
                cell.alignment = Alignment(horizontal="center", vertical="center")

    with pd.ExcelWriter(path_auditoria, engine='openpyxl') as writer:
        df_lista_auditoria.to_excel(writer, sheet_name='Tarefas_Auditoria', index=False)
        ws2 = writer.sheets['Tarefas_Auditoria']
        idx_inp = len(df_lista_auditoria.columns)
        for row in ws2.iter_rows(min_row=1):
            for cell in row:
                if cell.column == idx_inp:
                    cell.fill = f_inp
                    if cell.row == 1: cell.font = Font(bold=True, color="9C6500")
                cell.alignment = Alignment(horizontal="left", vertical="center")

    print(f"✅ Matriz (Ouro) salva em: {path_matriz}")
    print(f"✅ Auditoria (Prata) ATUALIZADA salva em: {path_auditoria}")

if __name__ == "__main__":
    gerar_relatorio_cadeiras()