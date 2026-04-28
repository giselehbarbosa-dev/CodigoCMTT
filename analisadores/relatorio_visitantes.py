"""
Módulo: Relatório de Visitantes e Lobby (Snapshot da IA)
Objetivo: Gerar o relatório de visitantes com base nos dados BRUTOS,
mantendo a rastreabilidade do que a máquina extraiu antes da auditoria.
"""

import pandas as pd
import os
from openpyxl.styles import PatternFill, Font, Alignment
from core import config_ambiente


def gerar_relatorio_visitantes():
    print("==========================================================")
    print("📊 INICIANDO GERAÇÃO DO RELATÓRIO DE VISITANTES (SNAPSHOT IA)")
    print("==========================================================\n")

    # Lê ESTRITAMENTE a base bruta (Bronze) para manter a rastreabilidade
    caminho_leitura = config_ambiente.CAMINHO_CSV_VISITANTES
    print(f"⏳ Lendo dados unificados brutos de: {caminho_leitura}")

    if not os.path.exists(caminho_leitura):
        print("❌ Arquivo unificado não encontrado! Rode o 'motor_extracao.py' primeiro.")
        return

    # Lê o arquivo OBT (One Big Table)
    df_visitantes = pd.read_csv(caminho_leitura, sep=';', encoding='utf-8-sig')

    if df_visitantes.empty:
        print("⚠️ O arquivo de visitantes está vazio.")
        return

    df_visitantes['Data'] = pd.to_datetime(df_visitantes['Data'], errors='coerce').dt.strftime('%d/%m/%Y')

    # SEPARANDO OS PÚBLICOS USANDO A COLUNA 'Tipo_Visitante'
    mask_ex_conselheiros = df_visitantes['Tipo_Visitante'].str.contains('Ex-Conselheiro', na=False, case=False)
    df_ex = df_visitantes[mask_ex_conselheiros].copy()
    df_ext_comuns = df_visitantes[~mask_ex_conselheiros].copy()

    # O arquivo gerado aqui é o Snapshot Histórico
    arquivo_saida = config_ambiente.CAMINHO_EXCEL_VISITANTES
    os.makedirs(os.path.dirname(arquivo_saida), exist_ok=True)

    print("📈 Processando rankings e exportando para Excel...")

    with pd.ExcelWriter(arquivo_saida, engine='openpyxl') as writer:

        # 1. ABA DE TODOS OS VISITANTES (A Base Completa)
        df_visitantes.to_excel(writer, sheet_name="Todos_os_Visitantes", index=False)

        # 2. RANKING DE EX-CONSELHEIROS
        if not df_ex.empty:
            rank_ex = df_ex.groupby(
                ['Nome_Oficial_Associado', 'Tipo_Visitante', 'Periodo_Mandato', 'Segmento']
            ).size().reset_index(name='Total_Presencas_Pos_Mandato')

            rank_ex = rank_ex.sort_values(by='Total_Presencas_Pos_Mandato', ascending=False)
            rank_ex.to_excel(writer, sheet_name="Ranking_Ex_Conselheiros", index=False)

        # 3. RANKING DE VISITANTES COMUNS (O Filtro do Lobby)
        if not df_ext_comuns.empty:
            df_ext_comuns['Nome_Oficial_Associado'] = df_ext_comuns['Nome_Oficial_Associado'].fillna("")

            rank_comum = df_ext_comuns.groupby(
                ['Nome_na_Ata', 'Tipo_Visitante', 'Nome_Oficial_Associado']
            ).size().reset_index(name='Total_Presencas')

            rank_comum = rank_comum.sort_values(by='Total_Presencas', ascending=False)

            # Limpa o ruído: mostra só quem foi mais de 1 vez
            rank_comum_limpo = rank_comum[rank_comum['Total_Presencas'] > 1]
            rank_comum_limpo.to_excel(writer, sheet_name="Ranking_Visitantes_Comuns", index=False)

        # ========================================================
        # 4. FORMATAÇÃO VISUAL (Padrão Camada Ouro)
        # ========================================================
        fill_cabecalho = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        font_cabecalho = Font(color="FFFFFF", bold=True)

        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            ws.freeze_panes = "A2"  # Congela a primeira linha

            # Pinta o cabeçalho de azul
            for cell in ws[1]:
                cell.fill = fill_cabecalho
                cell.font = font_cabecalho
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Alinha o conteúdo à esquerda para leitura de texto
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

    print("==========================================================")
    print(f"✅ RELATÓRIO SNAPSHOT DE VISITANTES PRONTO!")
    print(f"📂 Salvo em: {arquivo_saida}")
    print("==========================================================")


if __name__ == "__main__":
    gerar_relatorio_visitantes()