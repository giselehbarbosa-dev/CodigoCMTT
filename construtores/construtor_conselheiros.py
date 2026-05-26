"""
Módulo: Construtor de Conselheiros e Catálogo (Master Data Management)
Objetivo: Ler a base oficial, gerar o JSON e o Catálogo de Metadados Oficiais
replicando EXATAMENTE a estrutura de colunas original.
"""

import pandas as pd
import json
import os
import sys
from openpyxl.styles import PatternFill, Font, Alignment

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from core import config_ambiente

CAMINHO_EXCEL = config_ambiente.CAMINHO_EXCEL_MANDATOS
PASTA_SAIDA = config_ambiente.CAMINHO_CONFIGS
CAMINHO_CATALOGO = config_ambiente.CAMINHO_EXCEL_DICIONARIO

COL_FUNCAO, COL_SEGMENTO, COL_ORGAO = "FUNÇÃO", "SEGMENTO", "ÓRGÃO"
COL_PADRONIZADA, COL_NOME = "CADEIRA_PADRONIZADA", "NOME"
COL_GENERO, COL_CARGO_EXTRA = "GÊNERO", "CARGO_EXTRA"

TERMOS_IGNORAR = ["-", "NAO INDICADO", ""]
TERMOS_SEM_VOTO = ["CONVIDADO", "SECRETARIA EXECUTIVA", "APOIO ADMINISTRATIVO", "GESTÃO DO CONSELHO"]

# =====================================================================
# FUNÇÕES ORIGINAIS V65 (MATEMÁTICA INTACTA)
# =====================================================================

def limpar_texto(texto):
    if pd.isna(texto) or str(texto).strip().lower() == "nan" or texto == "":
        return ""
    return " ".join(str(texto).strip().split())

def verificar_se_vota(segmento, cadeira_padronizada):
    texto_analise = (str(segmento) + " " + str(cadeira_padronizada)).upper()
    for termo in TERMOS_SEM_VOTO:
        if termo in texto_analise: return False
    return True

def processar_aba(df, nome_aba):
    mandato_data = {"arquivo_origem": nome_aba, "cadeiras": []}
    cadeiras_map = {}

    for index, linha in df.iterrows():
        segmento = limpar_texto(linha.get(COL_SEGMENTO))
        orgao_original = limpar_texto(linha.get(COL_ORGAO))
        cadeira_padrao = limpar_texto(linha.get(COL_PADRONIZADA))
        funcao = limpar_texto(linha.get(COL_FUNCAO)).upper()
        genero = limpar_texto(linha.get(COL_GENERO)).upper()
        cargo_extra = limpar_texto(linha.get(COL_CARGO_EXTRA))

        nome_bruto = linha.get(COL_NOME, '')
        if pd.isna(nome_bruto) or str(nome_bruto).strip() == "":
            nome = "VAGO"
        else:
            nome = str(nome_bruto).strip()

        if not orgao_original: continue
        if not cadeira_padrao: cadeira_padrao = orgao_original.upper()

        chave = (segmento, cadeira_padrao)
        if chave not in cadeiras_map:
            cadeiras_map[chave] = {
                "segmento": segmento,
                "cadeira_padronizada": cadeira_padrao,
                "nomes_orgaos_originais": set(),
                "titulares": [],
                "suplentes": []
            }

        cadeiras_map[chave]["nomes_orgaos_originais"].add(orgao_original)

        if nome.upper() not in TERMOS_IGNORAR:
            membro_obj = {
                "nome": nome,
                "genero": genero if genero else None,
                "cargo_extra": cargo_extra if cargo_extra else None,
                "funcao_original": funcao
            }
            if "SUPLENTE" in funcao:
                cadeiras_map[chave]["suplentes"].append(membro_obj)
            else:
                cadeiras_map[chave]["titulares"].append(membro_obj)

    lista_final = []
    contagem_voto = 0
    for dados in cadeiras_map.values():
        lista_orgaos = sorted(list(dados["nomes_orgaos_originais"]))
        nome_exibicao = max(lista_orgaos, key=len) if lista_orgaos else dados["cadeira_padronizada"]

        objeto_cadeira = {
            "segmento": dados["segmento"],
            "nome_orgao_exibicao": nome_exibicao,
            "aliases_orgao": lista_orgaos,
            "cadeira_padronizada": dados["cadeira_padronizada"],
            "titulares": dados["titulares"],
            "suplentes": dados["suplentes"]
        }
        lista_final.append(objeto_cadeira)
        if verificar_se_vota(dados["segmento"], dados["cadeira_padronizada"]):
            contagem_voto += 1

    mandato_data["cadeiras"] = lista_final
    return mandato_data, contagem_voto

# =====================================================================
# GERAÇÃO DO CATÁLOGO DE METADADOS (ESTRUTURA ORIGINAL RESTAURADA)
# =====================================================================

def gerar_catalogo_metadados(lista_orgaos, lista_reps):
    print("\n📚 A gerar o Catálogo de Metadados Oficial (Governança)...")

    df_orgaos_bruto = pd.DataFrame(lista_orgaos)
    df_reps_bruto = pd.DataFrame(lista_reps)

    if df_reps_bruto.empty:
        print("⚠️ Nenhum dado encontrado para gerar o catálogo.")
        return

    # ABA 1: Evolução dos Órgãos (Pivotada)
    df_orgaos = df_orgaos_bruto.pivot_table(
        index=['Segmento', 'Cadeira Padronizada'],
        columns='Mandato',
        values='Orgao',
        aggfunc=lambda x: ' / '.join(x.dropna().unique())
    ).fillna('-').reset_index()

    # Ordena os mandatos
    colunas_fixas = ['Segmento', 'Cadeira Padronizada']
    colunas_mandatos = sorted([col for col in df_orgaos.columns if col not in colunas_fixas])
    df_orgaos = df_orgaos[colunas_fixas + colunas_mandatos]

    # ABA 2: Histórico de Conselheiros (Tabela Reta/Flat como no original)
    df_reps = df_reps_bruto.drop_duplicates()
    df_reps = df_reps.sort_values(
        by=['Segmento', 'Cadeira Padronizada', 'Periodo Mandato', 'Funcao'],
        ascending=[True, True, True, False] # Titular antes de Suplente
    )

    os.makedirs(os.path.dirname(CAMINHO_CATALOGO), exist_ok=True)
    with pd.ExcelWriter(CAMINHO_CATALOGO, engine='openpyxl') as writer:
        df_orgaos.to_excel(writer, sheet_name='Evolução_das_Secretarias', index=False)
        ws1 = writer.sheets['Evolução_das_Secretarias']
        ws1.freeze_panes = "C2"

        df_reps.to_excel(writer, sheet_name='Histórico_Conselheiros', index=False)
        ws2 = writer.sheets['Histórico_Conselheiros']
        ws2.freeze_panes = "A2" # Congela A2 como no original, não C2

        fill_cabecalho = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        font_cabecalho = Font(color="FFFFFF", bold=True)

        for ws in [ws1, ws2]:
            for cell in ws[1]:
                cell.fill = fill_cabecalho
                cell.font = font_cabecalho
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

    print(f"✅ Catálogo de Metadados salvo em: {CAMINHO_CATALOGO}")

# =====================================================================
# LOOP PRINCIPAL
# =====================================================================

def main():
    print("🚀 Iniciando Construtor de Conselheiros (V65 Intacta + Catálogo)")
    if not os.path.exists(PASTA_SAIDA): os.makedirs(PASTA_SAIDA)
    try:
        xls = pd.ExcelFile(CAMINHO_EXCEL)
    except Exception as e:
        print(f"❌ Erro ao abrir Excel: {e}"); return

    dados_orgaos_catalogo = []
    dados_reps_catalogo = []

    for nome_aba in xls.sheet_names:
        if nome_aba.lower() in ["listas", "ajuda", "config", "rascunho", "checklist", "exemplo"]: continue
        try:
            df = pd.read_excel(xls, sheet_name=nome_aba)
            df.columns = [str(c).strip().upper() for c in df.columns]
            if COL_ORGAO not in df.columns: continue

            dados, n_votos = processar_aba(df, nome_aba)

            caminho_json = os.path.join(PASTA_SAIDA, f"{nome_aba}.json")
            with open(caminho_json, 'w', encoding='utf-8') as f:
                json.dump(dados, f, indent=4, ensure_ascii=False)
            print(f"✅ JSON Gerado: {nome_aba}.json | Cadeiras com Voto: {n_votos}")

            # --- Mapeamento Idêntico ao CSV Original ---
            mandato_str = nome_aba.replace(" ", "_") # Ex: "2013ago_2014mai"

            for cadeira in dados['cadeiras']:
                # Popula dados para Aba 1
                dados_orgaos_catalogo.append({
                    'Segmento': cadeira['segmento'],
                    'Cadeira Padronizada': cadeira['cadeira_padronizada'],
                    'Mandato': mandato_str,
                    'Orgao': cadeira['nome_orgao_exibicao']
                })

                # Popula dados para Aba 2 (Linha por linha)
                for titular in cadeira['titulares']:
                    dados_reps_catalogo.append({
                        'Segmento': cadeira['segmento'],
                        'Cadeira Padronizada': cadeira['cadeira_padronizada'],
                        'Nome Conselheiro': titular['nome'],
                        'Funcao': 'TITULAR',
                        'Periodo Mandato': mandato_str
                    })

                for suplente in cadeira['suplentes']:
                    dados_reps_catalogo.append({
                        'Segmento': cadeira['segmento'],
                        'Cadeira Padronizada': cadeira['cadeira_padronizada'],
                        'Nome Conselheiro': suplente['nome'],
                        'Funcao': 'SUPLENTE',
                        'Periodo Mandato': mandato_str
                    })

        except Exception as e:
            print(f"❌ Erro na aba {nome_aba}: {e}")

    gerar_catalogo_metadados(dados_orgaos_catalogo, dados_reps_catalogo)

    print("==========================================================")
    print("🏆 PIPELINE MDM CONCLUÍDO COM LÓGICA E ESTRUTURA ORIGINAIS PRESERVADAS!")
    print("==========================================================")

if __name__ == "__main__":
    main()