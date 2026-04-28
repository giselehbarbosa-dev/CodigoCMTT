"""
Módulo: Construtor de Dicionário de Metadados (Data Governance)
Objetivo: Criar um dicionário histórico com duas abas:
1. Matriz de Órgãos: Mandatos como colunas e Órgãos como valores.
2. Histórico de Representantes: Titulares e Suplentes de cada cadeira por período.
"""

import os
import pandas as pd
from core import config_ambiente
from openpyxl.styles import PatternFill, Font, Alignment

def construir_dicionario():
    print("==========================================================")
    print("📚 INICIANDO A CONSTRUÇÃO DO CATÁLOGO DE METADADOS")
    print("==========================================================\n")

    caminho_prata = config_ambiente.CAMINHO_CSV_PRESENCA.replace('.csv', '_conferido.csv')
    caminho_base = caminho_prata if os.path.exists(caminho_prata) else config_ambiente.CAMINHO_CSV_PRESENCA

    if not os.path.exists(caminho_base):
        print("❌ ERRO: Nenhuma base de presença encontrada.")
        return

    print(f"⏳ Lendo base de dados: {os.path.basename(caminho_base)}")
    df = pd.read_csv(caminho_base, sep=';')

    # REMOVIDOS OS FILTROS RESTRITIVOS!
    # Agora os "CONVIDADOS" e as cadeiras com status "VAGO" vão aparecer perfeitamente
    # para refletir 100% a base_mandatos original.

    # ========================================================
    # ABA 1: MATRIZ PIVOTADA DE ÓRGÃOS (SECRETARIAS)
    # ========================================================
    print("⚙️ Pivotando a matriz de Órgãos por mandato...")
    df_orgaos = df.dropna(subset=['Segmento', 'Cadeira', 'Periodo_Mandato', 'Orgao']).copy()

    # Removemos as duplicatas pegando o último registro válido de cada mandato
    df_orgaos = df_orgaos.drop_duplicates(subset=['Segmento', 'Cadeira', 'Periodo_Mandato'], keep='last')

    # Renomeia para o formato pedido
    df_orgaos = df_orgaos.rename(columns={'Cadeira': 'Cadeira Padronizada'})

    # Pivot: Colunas viram os mandatos, Valores viram os órgãos
    matriz_orgaos = df_orgaos.pivot(
        index=['Segmento', 'Cadeira Padronizada'],
        columns='Periodo_Mandato',
        values='Orgao'
    ).reset_index()

    # Preenche vazios caso a cadeira não existisse em algum mandato
    matriz_orgaos = matriz_orgaos.fillna('-')

    # Ordena as colunas de mandatos alfabeticamente para ficarem na linha do tempo
    colunas_fixas = ['Segmento', 'Cadeira Padronizada']
    colunas_mandatos = sorted([col for col in matriz_orgaos.columns if col not in colunas_fixas])
    matriz_orgaos = matriz_orgaos[colunas_fixas + colunas_mandatos]

    # ========================================================
    # ABA 2: HISTÓRICO DE REPRESENTANTES
    # ========================================================
    print("⚙️ Construindo o histórico de Representantes...")
    df_reps = df.dropna(subset=['Segmento', 'Cadeira', 'Nome', 'Tipo', 'Periodo_Mandato']).copy()

    df_reps = df_reps.rename(columns={
        'Cadeira': 'Cadeira Padronizada',
        'Nome': 'Nome Conselheiro',
        'Tipo': 'Funcao',
        'Periodo_Mandato': 'Periodo Mandato'
    })

    # Extrai colunas de interesse e remove duplicatas
    colunas_reps = ['Segmento', 'Cadeira Padronizada', 'Nome Conselheiro', 'Funcao', 'Periodo Mandato']
    matriz_reps = df_reps[colunas_reps].drop_duplicates()

    # Ordena logicamente (Segmento -> Cadeira -> Mandato -> Titular 1º)
    matriz_reps = matriz_reps.sort_values(
        by=['Segmento', 'Cadeira Padronizada', 'Periodo Mandato', 'Funcao'],
        ascending=[True, True, True, False]
    )

    # ========================================================
    # EXPORTAÇÃO E FORMATAÇÃO (GOLD LAYER)
    # ========================================================
    caminho_saida = os.path.join(config_ambiente.CAMINHO_RELATORIOS, "Catálogo_de_Metadados_CMTT.xlsx")
    os.makedirs(os.path.dirname(caminho_saida), exist_ok=True)

    print("📈 Exportando para Excel...")
    with pd.ExcelWriter(caminho_saida, engine='openpyxl') as writer:
        # Aba 1
        matriz_orgaos.to_excel(writer, sheet_name='Evolução_das_Secretarias', index=False)
        ws1 = writer.sheets['Evolução_das_Secretarias']
        ws1.freeze_panes = "C2" # Congela a partir da coluna dos mandatos

        # Aba 2
        matriz_reps.to_excel(writer, sheet_name='Histórico_Conselheiros', index=False)
        ws2 = writer.sheets['Histórico_Conselheiros']
        ws2.freeze_panes = "A2"

        fill_cabecalho = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        font_cabecalho = Font(color="FFFFFF", bold=True)

        # Estilo Aba 1
        for cell in ws1[1]:
            cell.fill = fill_cabecalho
            cell.font = font_cabecalho
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in ws1.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # Estilo Aba 2
        for cell in ws2[1]:
            cell.fill = fill_cabecalho
            cell.font = font_cabecalho
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in ws2.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    print(f"🎉 Catálogo de Metadados gerado com sucesso!")
    print(f"📂 Salvo na Camada Ouro: {caminho_saida}")

if __name__ == "__main__":
    construir_dicionario()